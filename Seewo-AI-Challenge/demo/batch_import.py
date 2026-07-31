"""V2.0 Sprint 5 (5.12): 批量导入后端 — Excel/CSV 解析 + 校验 + 入库.

Endpoints (per API_CONTRACT.md):
  POST /api/admin/import-students           — 上传 + 预览校验
  POST /api/admin/import-students/confirm   — 确认入库
  GET  /api/admin/import-students/template  — 下载 Excel 模板

数据格式 (Excel/CSV 列):
  姓名 | 学号 | 班级名称 | 性别

校验规则:
  - 姓名：必填，1-20 字符
  - 学号：必填，唯一（不与已有学生重复）
  - 班级名称：必填，需在组织树中存在
  - 性别：可选，男/女
"""
from __future__ import annotations

import csv
import io
import json
from pathlib import Path

from flask import Blueprint, jsonify, request, send_file, g, session

from security import login_required, roles_required, data_scope, audit_log, get_current_user

bp = Blueprint("batch_import", __name__)

_DATA_DIR = Path(__file__).parent / "data"


def _parse_excel(file_bytes: bytes) -> list[dict]:
    """Parse .xlsx file into list of row dicts."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    result = []
    for i, row in enumerate(rows[1:], start=2):  # row 2+ (1-indexed, header is row 1)
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        record = {}
        for j, val in enumerate(row):
            if j < len(headers):
                record[headers[j]] = str(val).strip() if val is not None else ""
        record["_row"] = i
        result.append(record)
    return result


def _parse_csv(file_bytes: bytes) -> list[dict]:
    """Parse CSV bytes into list of row dicts."""
    text = file_bytes.decode("utf-8-sig")  # handle BOM
    reader = csv.DictReader(io.StringIO(text))
    result = []
    for i, row in enumerate(reader, start=2):
        cleaned = {k: (v.strip() if v else "") for k, v in row.items() if k}
        cleaned["_row"] = i
        result.append(cleaned)
    return result


def _normalize_row(row: dict) -> dict:
    """Normalize column names to standard fields."""
    # Handle various column name conventions
    name = row.get("姓名") or row.get("name") or row.get("Name") or ""
    student_no = row.get("学号") or row.get("student_no") or row.get("student_id") or ""
    class_name = row.get("班级名称") or row.get("班级") or row.get("class_name") or row.get("class") or ""
    gender = row.get("性别") or row.get("gender") or row.get("sex") or ""

    return {
        "name": name.strip(),
        "student_no": student_no.strip(),
        "class_name": class_name.strip(),
        "gender": gender.strip(),
        "_row": row.get("_row", 0),
    }


def _validate_rows(rows: list[dict], existing_student_nos: set, existing_class_names: set) -> tuple[list, list]:
    """Validate normalized rows. Returns (valid_rows, errors)."""
    valid = []
    errors = []
    seen_nos = set()

    for row in rows:
        nr = _normalize_row(row)
        row_num = nr["_row"]
        name = nr["name"]
        student_no = nr["student_no"]
        class_name = nr["class_name"]

        # Required field checks
        if not name:
            errors.append({"row": row_num, "reason": "缺少姓名"})
            continue
        if len(name) > 20:
            errors.append({"row": row_num, "name": name, "reason": "姓名超过20字符"})
            continue
        if not student_no:
            errors.append({"row": row_num, "name": name, "reason": "缺少学号"})
            continue
        if not class_name:
            errors.append({"row": row_num, "name": name, "reason": "缺少班级名称"})
            continue

        # Duplicate check (within file)
        if student_no in seen_nos:
            errors.append({"row": row_num, "name": name, "reason": f"学号 {student_no} 在文件内重复"})
            continue
        seen_nos.add(student_no)

        # Duplicate check (against existing)
        if student_no in existing_student_nos:
            errors.append({"row": row_num, "name": name, "reason": f"学号 {student_no} 已存在"})
            continue

        # Class existence check
        if class_name not in existing_class_names:
            errors.append({"row": row_num, "name": name, "reason": f"班级 '{class_name}' 不存在"})
            continue

        # Gender validation (optional)
        if nr["gender"] and nr["gender"] not in ("男", "女", "M", "F", "male", "female"):
            errors.append({"row": row_num, "name": name, "reason": f"性别 '{nr['gender']}' 无效"})
            continue

        valid.append(nr)

    return valid, errors


def _get_existing_student_nos() -> set:
    """Get set of existing student IDs from students.json."""
    try:
        from engine.grader import load_json
        students = load_json("students.json").get("students", [])
        return {s.get("student_id", s.get("id", "")) for s in students}
    except Exception:
        return set()


def _get_existing_class_names() -> set:
    """Get set of existing class names from organization.json or students.json."""
    org_file = _DATA_DIR / "organization.json"
    if org_file.exists():
        with open(org_file, "r", encoding="utf-8") as f:
            org = json.load(f)
        return {c.get("name", "") for c in org.get("classes", [])}
    # Fallback: derive from students.json
    try:
        from engine.grader import load_json
        students = load_json("students.json").get("students", [])
        return {s.get("class_name", "") for s in students if s.get("class_name")}
    except Exception:
        return set()


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@bp.route("/api/admin/import-students", methods=["POST"])
@login_required
@roles_required("admin", "head")
@data_scope()
def import_students_preview():
    """Upload Excel/CSV and return preview with validation results."""
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "no file uploaded"}), 400

    file = request.files["file"]
    filename = file.filename or ""
    file_bytes = file.read()

    # Parse based on extension
    if filename.endswith(".xlsx"):
        rows = _parse_excel(file_bytes)
    elif filename.endswith(".csv"):
        rows = _parse_csv(file_bytes)
    else:
        return jsonify({"ok": False, "error": "unsupported file format (use .xlsx or .csv)"}), 400

    # Validate
    existing_nos = _get_existing_student_nos()
    existing_classes = _get_existing_class_names()
    valid, errors = _validate_rows(rows, existing_nos, existing_classes)

    # Build preview
    preview = [
        {
            "row": r["_row"],
            "name": r["name"],
            "student_no": r["student_no"],
            "class_name": r["class_name"],
            "valid": True,
        }
        for r in valid
    ]

    return jsonify({
        "ok": True,
        "data": {
            "total": len(rows),
            "valid": len(valid),
            "invalid": len(errors),
            "preview": preview[:100],  # limit preview to first 100
            "errors": errors,
        }
    })


@bp.route("/api/admin/import-students/confirm", methods=["POST"])
@login_required
@roles_required("admin", "head")
@data_scope()
def import_students_confirm():
    """Confirm import: write validated rows into students.json."""
    body = request.get_json(silent=True) or {}
    rows = body.get("rows", [])
    class_id = body.get("class_id")

    if not rows:
        return jsonify({"ok": False, "error": "no rows to import"}), 400

    # Load existing students
    students_file = _DATA_DIR / "students.json"
    if students_file.exists():
        with open(students_file, "r", encoding="utf-8") as f:
            students_data = json.load(f)
    else:
        students_data = {"students": []}

    existing_nos = {s.get("student_id", s.get("id", "")) for s in students_data["students"]}
    imported = 0

    for row in rows:
        student_no = row.get("student_no", "")
        if student_no in existing_nos:
            continue  # skip duplicates

        # Generate student_id
        max_num = 0
        for s in students_data["students"]:
            sid = s.get("student_id", "")
            if sid.startswith("s"):
                try:
                    num = int(sid[1:])
                    max_num = max(max_num, num)
                except ValueError:
                    pass

        new_student = {
            "student_id": f"s{max_num + 1:02d}",
            "name": row.get("name", ""),
            "student_no": student_no,
            "class_name": row.get("class_name", ""),
            "class_id": class_id,
            "gender": row.get("gender", ""),
            "level": "C",  # default level
        }
        students_data["students"].append(new_student)
        existing_nos.add(student_no)
        imported += 1

    # Save
    tmp = students_file.with_suffix(".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(students_data, f, ensure_ascii=False, indent=2)
    tmp.replace(students_file)

    school_id = getattr(g, "school_id", 1)
    audit_log("students_imported", resource="students",
              imported=imported, total_rows=len(rows), school_id=school_id)
    return jsonify({"ok": True, "data": {"imported": imported}})


@bp.route("/api/admin/import-students/template")
@login_required
@roles_required("admin", "head")
def import_students_template():
    """Download Excel template for student import."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生导入模板"
    ws.append(["姓名", "学号", "班级名称", "性别"])
    # Example row
    ws.append(["张三", "2026001", "高二(1)班", "男"])
    ws.append(["李四", "2026002", "高二(1)班", "女"])

    # Set column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 8

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="学生导入模板.xlsx",
    )
